import openpyxl
from openpyxl.styles import Font
from database.db import get_all_registrations, get_registrations_by_event

async def generate_excel_for_all_events(filename: str):
    from openpyxl import Workbook
    from database.db import get_all_registrations

    registrations = await get_all_registrations()

    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    ws.append(["Мероприятие", "Имя", "Фамилия", "Email", "Телефон", "Telegram"])

    for r in registrations:
        ws.append([r["event"], r["name"], r["surname"], r["email"], r["phone"], r["telegram"]])

    wb.save(filename)


def generate_excel_for_event(event_title, filename="event.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = event_title

    headers = ["ФИО", "Email", "Телефон", "Telegram"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    registrations = get_registrations_by_event_sync(event_title)

    for r in registrations:
        ws.append([
            f"{r['name']} {r['surname']}",
            r['email'],
            r['phone'],
            r['telegram']
        ])

    wb.save(filename)

# Синхронные обёртки для вызова из обычного кода
def get_all_registrations_sync():
    import asyncio
    from database.db import get_all_registrations
    return asyncio.run(get_all_registrations())

def get_registrations_by_event_sync(event_title):
    import asyncio
    from database.db import get_registrations_by_event
    return asyncio.run(get_registrations_by_event(event_title))
